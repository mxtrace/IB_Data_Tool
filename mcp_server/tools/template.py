"""
template.py — Excel 模板填充工具
根据模板类型（AMS/ENS/ASI）填充进仓数量和收发通信息。

单元格映射（已确认）：
  AMS (Sheet "AMS"):
    Shipper:     B7=Company, B8=Address, B9=City, D9=State, B10=Zip, D10=Country
    Consignee:   B12=Company, B13=Address, B14=City, D14=State, B15=Zip, D15=Country
    Notify:      B17=Company, B18=Address, B19=City, D19=State, B20=Zip, D20=Country
    Quantity:    C23=件数, D23=CARTON, E23=重量(kg), F23=体积(CBM)

  ENS (Sheet "ENS"):
    Shipper:     B7=Company, B8=Address, B9=City, D9=State, B10=Zip, D10=Country
    Consignee:   B12=Company, B13=Address, B14=City, D14=State, B15=Zip, D15=Country
    Buyer:       G12=Company, G13=Address, G14=City, I14=State, G15=Zip, I15=Country
    Notify:      B17=Company, B18=Address, B19=City, D19=State, B20=Zip, D20=Country
    Quantity:    C23=件数, D23=CARTON, E23=重量(kg), F23=体积(CBM)

  ASI: 仅填充 Quantity 行（C23/D23/E23/F23），其余由 OC 预填。
"""
from __future__ import annotations

import shutil
from pathlib import Path

from tools.config import get_config

DATA_DIR = Path(__file__).parent.parent / "data"
OUTPUT_DIR = DATA_DIR / "output"


def fill_template(
    al0: str,
    template_type: str,
    ib_row: dict,
    parties_data: dict = None,
    asi_file_path: str = None,
) -> dict:
    """
    填充模板。

    Args:
        al0: 订单 ID
        template_type: "AMS" | "ENS" | "ASI"
        ib_row: 进仓数据行 {received_cartons, received_volume, received_weight, flipped_fc, ...}
        parties_data: Fetch Session 返回的 Parties 地址数据
                      格式: {contactId: {company_name, address_line1, city, ...}}
        asi_file_path: ASI 下载文件路径（CDA时必须提供）

    Returns:
        {success, output_file, template_type, error}
    """
    config = get_config()
    if "error" in config:
        return {"success": False, "error": config["error"]}

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    try:
        if template_type == "ASI":
            return _fill_asi(al0, ib_row, asi_file_path)
        elif template_type in ("AMS", "ENS"):
            return _fill_ams_ens(al0, template_type, ib_row, parties_data, config)
        else:
            return {"success": False, "error": f"未知模板类型: {template_type}"}
    except Exception as e:
        return {"success": False, "error": str(e)}


# ══════════════════════════════════════════════════════════════════════
# ASI 填充
# ══════════════════════════════════════════════════════════════════════

def _fill_asi(al0: str, ib_row: dict, asi_file_path: str) -> dict:
    """ASI 模板：仅填充数量行。"""
    if not asi_file_path:
        return {"success": False, "error": "ASI 文件路径为空"}

    src = Path(asi_file_path)
    if not src.exists():
        return {"success": False, "error": f"ASI 文件不存在: {asi_file_path}"}

    output_file = OUTPUT_DIR / f"{al0}_ASI{src.suffix}"
    shutil.copy2(src, output_file)

    _write_quantities(output_file, ib_row)

    return {"success": True, "output_file": str(output_file), "template_type": "ASI"}


# ══════════════════════════════════════════════════════════════════════
# AMS / ENS 填充
# ══════════════════════════════════════════════════════════════════════

def _fill_ams_ens(al0: str, ttype: str, ib_row: dict, parties_data: dict, config: dict) -> dict:
    """AMS/ENS 模板：填充数量 + 收发通。"""

    # ── 1. FC 地址匹配（强校验，失败即 BLOCKER）──
    flipped_fc = ib_row.get("flipped_fc", "")
    fc_address = _match_fc_address(flipped_fc, config.get("fc_address_path", ""))
    if not fc_address:
        return {
            "success": False,
            "error": f"[BLOCKER] flipped_fc={flipped_fc} 在 FC_Address.xlsx 中无匹配，禁止兜底",
        }

    # ── 2. 复制模板 ──
    if ttype == "AMS":
        template_path = Path(config.get("ams_template_path", "data/AMS_template.xlsx"))
    else:
        template_path = Path(config.get("ens_template_path", "data/ENS_template.xlsx"))

    if not template_path.is_absolute():
        template_path = DATA_DIR.parent / template_path

    if not template_path.exists():
        return {"success": False, "error": f"模板文件不存在: {template_path}"}

    output_file = OUTPUT_DIR / f"{al0}_{ttype}.xlsx"
    shutil.copy2(template_path, output_file)

    # ── 3. 填充数量（Row 23）──
    _write_quantities(output_file, ib_row)

    # ── 4. 填充收发通 ──
    if parties_data:
        _write_parties(output_file, ttype, ib_row, parties_data, fc_address)

    return {"success": True, "output_file": str(output_file), "template_type": ttype}


# ══════════════════════════════════════════════════════════════════════
# 数量填充
# ══════════════════════════════════════════════════════════════════════

def _write_quantities(file_path: Path, ib_row: dict) -> None:
    """
    Row 23: C23=件数, D23=CARTON, E23=重量(kg), F23=体积(CBM)
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(file_path))
    ws = wb.active

    ws["C23"] = ib_row.get("received_cartons", 0)
    ws["D23"] = "CARTON"
    ws["E23"] = ib_row.get("received_weight", 0)   # E=重量
    ws["F23"] = ib_row.get("received_volume", 0)    # F=体积

    wb.save(str(file_path))
    wb.close()


# ══════════════════════════════════════════════════════════════════════
# 收发通填充
# ══════════════════════════════════════════════════════════════════════

def _write_parties(
    file_path: Path,
    ttype: str,
    ib_row: dict,
    parties_data: dict,
    fc_address: dict,
) -> None:
    """
    填充 Shipper / Consignee / Notify Party / Buyer(ENS)。

    数据来源：
    - Shipper: API Parties 返回（shipperContactId 对应条目）
    - Consignee Company: API company + "\\nC/O FBA"
    - Consignee Address: FC_Address 表（强制，不取API地址）
    - Notify Party: Company 取 API，Address 同 Consignee（FC_Address）
    - Buyer (ENS only): API importerParty 对应条目
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(file_path))
    ws = wb.active

    # 从 parties_data 中按 contactId 取各方数据
    # parties_data 格式: {contactId: {companyName, addressLine1, city, ...}}
    # 需要从 booking_results 知道哪个 contactId 对应哪个角色
    # 这里假设 parties_data 已经按角色组织好:
    #   parties_data["shipper"] = {...}
    #   parties_data["consignee"] = {...}
    #   parties_data["notify"] = {...}
    #   parties_data["buyer"] = {...}  (ENS only)

    shipper = parties_data.get("shipper", {})
    consignee_api = parties_data.get("consignee", {})
    notify_api = parties_data.get("notify", {})
    buyer = parties_data.get("buyer", {})

    # ── Shipper (Row 7-10): 完全取 API ──
    ws["B7"] = shipper.get("companyName", "")
    ws["B8"] = _join_address_lines(shipper)
    ws["B9"] = shipper.get("city", "")
    ws["D9"] = shipper.get("stateOrRegion", "")
    ws["B10"] = shipper.get("postalCode", "")
    ws["D10"] = shipper.get("countryCode", "")

    # ── Consignee (Row 12-15): Company取API+C/O FBA, 地址取FC_Address ──
    consignee_company = consignee_api.get("companyName", "")
    if consignee_company and "C/O FBA" not in consignee_company.upper():
        consignee_company = f"{consignee_company} c/o FBA"
    ws["B12"] = consignee_company
    ws["B13"] = fc_address.get("address", "")
    ws["B14"] = fc_address.get("city", "")
    ws["D14"] = fc_address.get("state", "")
    ws["B15"] = fc_address.get("postal_code", "")
    ws["D15"] = fc_address.get("country", "")

    # ── Notify Party (Row 17-20): Company取API, 地址同Consignee(FC_Address) ──
    ws["B17"] = notify_api.get("companyName", "")
    ws["B18"] = fc_address.get("address", "")
    ws["B19"] = fc_address.get("city", "")
    ws["D19"] = fc_address.get("state", "")
    ws["B20"] = fc_address.get("postal_code", "")
    ws["D20"] = fc_address.get("country", "")

    # ── Buyer (ENS only, F/G 列 Row 12-15) ──
    if ttype == "ENS" and buyer:
        ws["G12"] = buyer.get("companyName", "")
        ws["G13"] = _join_address_lines(buyer)
        ws["G14"] = buyer.get("city", "")
        ws["I14"] = buyer.get("stateOrRegion", "")
        ws["G15"] = buyer.get("postalCode", "")
        ws["I15"] = buyer.get("countryCode", "")

    wb.save(str(file_path))
    wb.close()


def _join_address_lines(addr: dict) -> str:
    """拼接 addressLine1/2/3 为单行地址。"""
    parts = []
    for key in ("addressLine1", "addressLine2", "addressLine3"):
        val = addr.get(key)
        if val:
            parts.append(val.strip())
    return ", ".join(parts)


# ══════════════════════════════════════════════════════════════════════
# FC 地址匹配
# ══════════════════════════════════════════════════════════════════════

def _match_fc_address(flipped_fc: str, fc_path: str) -> dict | None:
    """
    在 FC_Address.xlsx 中匹配 FC 代码。
    匹配 A 列（FC），返回 B~H 列数据。
    匹配失败返回 None → BLOCKER。

    FC_Address 表结构:
      A=FC, B=POD, C=COMPANY_NAME, D=ADDRESS, E=CITY, F=STATES, G=POSTAL_CODE, H=COUNTRY
    """
    if not flipped_fc:
        return None

    fc_file = Path(fc_path)
    if not fc_file.is_absolute():
        fc_file = Path(__file__).parent.parent / fc_path
    if not fc_file.exists():
        return None

    import openpyxl
    wb = openpyxl.load_workbook(str(fc_file), read_only=True, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        fc_code = str(row[0] or "").strip()
        if fc_code.upper() == flipped_fc.upper():
            result = {
                "fc": fc_code,
                "pod": str(row[1] or "").strip(),
                "company_name": str(row[2] or "").strip(),
                "address": str(row[3] or "").strip(),
                "city": str(row[4] or "").strip(),
                "state": str(row[5] or "").strip(),
                "postal_code": str(row[6] or "").strip() if row[6] else "",
                "country": str(row[7] or "").strip(),
            }
            wb.close()
            return result

    wb.close()
    return None
