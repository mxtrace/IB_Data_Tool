"""
test_step345.py - Step 3/4/5 Integration Test
Uses parsed data from Step2 to test template fill + email generation.
Usage: python test_step345.py AL0-T33GH6PBKWGDW
"""
from __future__ import annotations
import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

from pathlib import Path
from core.input_zone_parser import parse_input_zone, InputZoneData
from steps.step3_template_select import select_template
from steps.step4_fill_template import fill_template, FillResult
from steps.step5_email import generate_email, EmailResult
from core.config import AppConfig

BASE_DIR = Path(__file__).parent


def main():
    al0 = sys.argv[1] if len(sys.argv) > 1 else 'AL0-T33GH6PBKWGDW'
    print(f'\n{"="*60}')
    print(f'  Step 3/4/5 Integration Test')
    print(f'  AL0: {al0}')
    print(f'{"="*60}\n')

    # Load parsed text from debug file
    debug_file = BASE_DIR / 'debug' / f'{al0}_raw.txt'
    if not debug_file.exists():
        print(f'ERROR: Run test_integration.py first to generate debug/{al0}_raw.txt')
        sys.exit(1)

    raw_text = debug_file.read_text(encoding='utf-8')
    input_zone = parse_input_zone(raw_text)
    print(f'[loaded] ODM={input_zone.odm_booking}, CDA={input_zone.cda_booking}')
    print(f'[loaded] Shipper={input_zone.shipper.company}')

    # Simulate ib_row (normally from Pending List)
    ib_row = {
        'al0': al0,
        'hbl_number': 'TEST-HBL-001',
        'shipper_company_name': input_zone.shipper.company,
        'pol': 'CNSZX',
        'pod': 'USOAK',
        'bc_login': 'miaoyua',
        'flipped_fc': 'TCY2',
        'received_cartons': 42,
        'received_volume': 3.56,
        'received_weight': 580.5,
    }

    # === Step 3: Template Selection ===
    print(f'\n[Step 3] Template Selection...')
    ttype = select_template(ib_row['pod'])
    print(f'  POD={ib_row["pod"]} -> Template type: {ttype}')

    # === Step 4: Fill Template ===
    print(f'\n[Step 4] Fill Template...')
    config = AppConfig(selected_logins=['miaoyua'])
    fill_result = fill_template(
        al0=al0,
        template_type=ttype,
        ib_row=ib_row,
        input_zone=input_zone,
        base_dir=BASE_DIR,
        asi_file=None,
    )
    if fill_result.error:
        print(f'  FAIL: {fill_result.error}')
        return
    print(f'  OK Output: {fill_result.output_file}')

    # Verify output file content
    import openpyxl
    wb = openpyxl.load_workbook(fill_result.output_file, data_only=True)
    ws = wb.active
    print(f'  Verify B7 (Shipper Company): {ws["B7"].value}')
    print(f'  Verify B8 (Shipper Address): {ws["B8"].value}')
    print(f'  Verify B12 (Consignee):      {ws["B12"].value}')
    print(f'  Verify B13 (Cons Address):   {ws["B13"].value}')
    print(f'  Verify C23 (Cartons):        {ws["C23"].value}')
    print(f'  Verify E23 (Weight):         {ws["E23"].value}')
    print(f'  Verify F23 (Volume):         {ws["F23"].value}')
    wb.close()

    # === Step 5: Generate Email ===
    print(f'\n[Step 5] Generate Email...')
    email_result = generate_email(
        al0=al0,
        ib_row=ib_row,
        input_zone=input_zone,
        attachment_path=fill_result.output_file,
        config=config,
        base_dir=BASE_DIR,
    )
    if email_result.error:
        print(f'  FAIL: {email_result.error}')
    else:
        print(f'  OK Email window displayed!')

    print(f'\n{"="*60}')
    print(f'  Test complete!')
    print(f'{"="*60}')


if __name__ == '__main__':
    main()
