"""
step1_sync.py — 读取进仓数据 → 同步到 Pending List → 取批次
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from core.config import AppConfig
from core.batch_controller import BatchData, TicketResult
from core.excel_helper import open_workbook, open_workbook_write
from core.logger import audit, log_info

# 进仓数据列映射（0-indexed）
COL = {
    "booking_id": 0,
    "hbl_number": 1,
    "shipper_company_name": 2,
    "origin": 5,
    "pod": 6,
    "bc": 8,
    "flipped_fc": 10,
    "received_cartons": 13,
    "received_volume": 14,
    "received_weight": 15,
    "container_count_by_dimension": 18,
    "check": 19,
}

HEADERS = [
    "booking_id", "hbl_number", "shipper_company_name", "fba_no",
    "pol_palletization", "origin", "pod", "operator", "bc", "fc",
    "flipped_fc", "booking_completed_date", "unload_completed_date",
    "received_cartons", "received_volume", "received_weight",
    "total_pallet_container_count", "pallet_dimension",
    "container_count_by_dimension", "check", "dataSource", "IB_Count",
]


def sync_to_pending_list(config: AppConfig, base_dir: Path) -> dict:
    """
    Step 1.1~1.3: 读本地 PendingList → 过滤 → 追加到工具 Pending List
    数据源：Desktop/Mars_LCL_Package/BookingFilePack/BCFile/PendingList/
    """
    import os

    # 1.1 定位本地 PendingList 目录（最新 .xlsx）
    username = os.environ.get("USERNAME", "")
    pending_src_dir = Path("C:/Users") / username / "Desktop" / "Mars_LCL_Package" / "BookingFilePack" / "BCFile" / "PendingList"
    if not pending_src_dir.exists():
        raise RuntimeError(f"PendingList 目录不存在：{pending_src_dir}")

    xlsx_files = sorted(
        [f for f in pending_src_dir.glob("*.xlsx") if not f.name.startswith("~$")],
        key=lambda f: f.stat().st_mtime, reverse=True,
    )
    if not xlsx_files:
        raise RuntimeError(f"PendingList 目录下未找到 .xlsx 文件：{pending_src_dir}")

    source_file = xlsx_files[0]
    log_info(f"Step1: 读取本地 PendingList {source_file.name}")

    # 1.2 读取 + 过滤（按 bc_login）
    with open_workbook(source_file, read_only=True, data_only=True) as wb:
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

    selected_set = {login.lower().strip() for login in config.selected_logins}
    filtered = []
    for row in rows:
        bc_val = str(row[COL["bc"]] or "").strip().lower()
        if bc_val in selected_set:
            filtered.append(row)

    log_info(f"Step1: PendingList {len(rows)} 行，过滤后 {len(filtered)} 行")

    # 1.3 去重：仅排除工具 Pending List 中尚未处理的（check列为空）
    pl_path = _get_pending_list_path(base_dir, config)
    if filtered:
        pending_al0s = _load_pending_al0s(pl_path)

        deduped = [r for r in filtered if str(r[COL["booking_id"]] or "").strip() not in pending_al0s]
        skipped = len(filtered) - len(deduped)
        if skipped:
            log_info(f"Step1: 去重跳过 {skipped} 行（当前批次中已存在）")
        filtered = deduped

    # 1.4 追加到工具 Pending List
    if filtered:
        _append_to_pending_list(pl_path, filtered)

    return {"source": str(source_file), "appended": len(filtered)}


def load_batch(config: AppConfig, base_dir: Path) -> BatchData:
    """
    Step 1.4: 从 Pending List 取 T列=空 的前 batch_size 条
    """
    pl_path = _get_pending_list_path(base_dir, config)
    if not pl_path.exists():
        return BatchData(al0_list=[], rows={}, row_indices={})

    with open_workbook(pl_path, read_only=False, data_only=True) as wb:
        ws = wb.active

        al0_list = []
        rows_dict = {}
        row_indices = {}
        total_pending = 0

        for row_idx in range(2, ws.max_row + 1):
            values = [ws.cell(row=row_idx, column=c).value for c in range(1, 23)]
            check_val = values[COL["check"]] if len(values) > COL["check"] else None

            if check_val and str(check_val).strip():
                continue  # 已处理

            al0 = str(values[COL["booking_id"]] or "").strip()
            if not al0:
                continue

            total_pending += 1

            if len(al0_list) >= config.batch_size:
                continue  # 继续计数但不加入本批次

            al0_list.append(al0)
            row_indices[al0] = row_idx
            rows_dict[al0] = {
                "al0": al0,
                "hbl_number": str(values[1] or "").strip(),
                "shipper_company_name": str(values[2] or "").strip(),
                "pol": str(values[5] or "").strip(),
                "pod": str(values[6] or "").strip(),
                "bc_login": str(values[8] or "").strip(),
                "flipped_fc": str(values[10] or "").strip(),
                "received_cartons": _to_float(values[13]),
                "received_volume": _to_float(values[14]),
                "received_weight": _to_float(values[15]),
            }

    return BatchData(al0_list=al0_list, rows=rows_dict, row_indices=row_indices, total_pending=total_pending)


def write_back_results(batch: BatchData, results: list[TicketResult], base_dir: Path):
    """批次完成后统一回写 Pending List（由 WAL 保障不重复）"""
    from core.config import AppConfig
    pl_path = _find_pending_list_file(base_dir)
    if not pl_path:
        return

    with open_workbook_write(pl_path) as wb:
        ws = wb.active
        for r in results:
            row_idx = batch.row_indices.get(r.al0)
            if not row_idx:
                continue
            if r.status == "success":
                ws.cell(row=row_idx, column=COL["check"] + 1, value="Yes")
            elif r.status == "odm":
                ws.cell(row=row_idx, column=COL["container_count_by_dimension"] + 1, value="ODM")
                ws.cell(row=row_idx, column=COL["check"] + 1, value="Yes")


def write_single_ticket(base_dir: Path, al0: str, status: str):
    """
    WAL 恢复用：单票回写。
    扫描 Pending List 找到 al0 对应行，写入状态。
    """
    pl_path = _find_pending_list_file(base_dir)
    if not pl_path:
        return

    with open_workbook_write(pl_path) as wb:
        ws = wb.active
        for row_idx in range(2, ws.max_row + 1):
            cell_al0 = str(ws.cell(row=row_idx, column=1).value or "").strip()
            cell_check = ws.cell(row=row_idx, column=COL["check"] + 1).value
            if cell_al0 == al0 and not cell_check:
                if status == "success":
                    ws.cell(row=row_idx, column=COL["check"] + 1, value="Yes")
                elif status == "odm":
                    ws.cell(row=row_idx, column=COL["container_count_by_dimension"] + 1, value="ODM")
                    ws.cell(row=row_idx, column=COL["check"] + 1, value="Yes")
                break


# ── 去重辅助 ──

def _load_existing_al0s(pl_path: Path) -> set:
    """从 Pending List 中提取所有已有 AL0"""
    if not pl_path.exists():
        return set()
    try:
        with open_workbook(pl_path, read_only=True, data_only=True) as wb:
            ws = wb.active
            al0s = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                al0 = str(row[0] or "").strip()
                if al0:
                    al0s.add(al0)
            return al0s
    except Exception:
        return set()


def _load_pending_al0s(pl_path: Path) -> set:
    """从 Pending List 中提取尚未处理的 AL0（check列为空）"""
    if not pl_path.exists():
        return set()
    try:
        with open_workbook(pl_path, read_only=True, data_only=True) as wb:
            ws = wb.active
            al0s = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                al0 = str(row[0] or "").strip()
                check = str(row[COL["check"]] or "").strip() if len(row) > COL["check"] else ""
                if al0 and not check:
                    al0s.add(al0)
            return al0s
    except Exception:
        return set()


def _load_history(base_dir: Path) -> set:
    """从 processed.jsonl 中加载历史已处理 AL0"""
    import json
    history_file = base_dir / "processed.jsonl"
    if not history_file.exists():
        return set()
    al0s = set()
    for line in history_file.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line:
            try:
                entry = json.loads(line)
                al0s.add(entry["al0"])
            except (json.JSONDecodeError, KeyError):
                continue
    return al0s


def append_to_history(base_dir: Path, al0: str, status: str):
    """记录已处理的 AL0 到 history（防重复发邮件）"""
    import json
    from datetime import datetime
    history_file = base_dir / "processed.jsonl"
    entry = {"al0": al0, "status": status, "ts": datetime.now().isoformat()}
    with open(history_file, "a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")


# ── 内部辅助 ──

def _get_pending_list_path(base_dir: Path, config: AppConfig) -> Path:
    pl_dir = base_dir / config.pending_list_dir
    pl_dir.mkdir(parents=True, exist_ok=True)
    existing = sorted(pl_dir.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    if existing:
        return existing[0]
    return pl_dir / "pending_list.xlsx"


def _find_pending_list_file(base_dir: Path) -> Path | None:
    pl_dir = base_dir / "Pending list"
    if not pl_dir.exists():
        return None
    existing = sorted(pl_dir.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    return existing[0] if existing else None


def _append_to_pending_list(pl_path: Path, rows: list):
    """追加行到 Pending List"""
    if pl_path.exists():
        wb = openpyxl.load_workbook(str(pl_path))
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(HEADERS)

    for row in rows:
        row_list = list(row)[:22] if len(row) >= 22 else list(row) + [None] * (22 - len(row))
        row_list[18] = None  # S列清空
        row_list[19] = None  # T列清空
        ws.append(row_list)

    wb.save(str(pl_path))
    wb.close()


def _to_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


